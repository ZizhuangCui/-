#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
上海政府采购公告爬虫 v15 - 格式优化版
优化报名时间提取，优化 Excel 列宽和格式
"""

import asyncio
from playwright.async_api import async_playwright
import requests
import pandas as pd
from datetime import datetime
import time
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 弱电工程关键词
WEAK_ELECTRIC_KEYWORDS = [
    '安防', '监控', '视频监控', '门禁', '入侵报警', '周界', '人脸识别', '智能门禁',
    '道闸', '人行通道', '车辆测速', '一键报警', '升降柱', '电子巡更', '对讲', '电子警察',
    '安检', '防盗', '保安', '安全系统', '智能化', '弱电', '智能建筑', '楼宇自控', 
    '智能照明', '能源管理', '智能', '综合布线', '网络工程', '无线覆盖', 'WiFi', 
    '通信系统', '光纤', '机房', '网络', '布线', '局域网', '广域网', '交换机', '路由器',
    '会议系统', '广播系统', '多媒体', '音响', '投影', 'LED屏', '显示屏', '视听',
    '扩声', '舞台', '演播', '录播', '一卡通', '停车管理', '停车场', '车位',
    '数据中心', '机房工程', 'UPS', '精密空调', '动环监控', '服务器', '存储',
    '电池', '电源', '配电', '系统集成', '信息化', '数字化', '云平台', '物联网', 
    'IoT', '大数据', '系统建设', '系统升级', '系统改造', '系统开发', '系统维护', 
    '系统运维', '智慧校园', '智慧医院', '智慧城市', '智慧园区', '智慧楼宇', 
    '智慧交通', '智慧', '政务云', '数智', '电子', '科技', '技术', '自动化', 
    '控制', '软件', '硬件', 'IT', '信息',
]

API_URL = "https://www.zfcg.sh.gov.cn/portal/category"
BASE_URL = "https://www.zfcg.sh.gov.cn"
CATEGORY_CODE = "ZcyAnnouncement2"
PARENT_ID = "137027"
PAGE_SIZE = 15

HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Content-Type": "application/json;charset=UTF-8",
    "Referer": "https://www.zfcg.sh.gov.cn/",
    "Origin": "https://www.zfcg.sh.gov.cn",
}

def is_weak_electric(project_name):
    if pd.isna(project_name):
        return False, []
    project_name = str(project_name)
    matched = [k for k in WEAK_ELECTRIC_KEYWORDS if k in project_name]
    return len(matched) > 0, matched

def format_date_from_ts(v):
    if v is None or v == "":
        return ""
    if isinstance(v, int):
        s = str(v)
        if len(s) == 13:
            return datetime.fromtimestamp(v / 1000).strftime("%Y-%m-%d")
        if len(s) == 10:
            return datetime.fromtimestamp(v).strftime("%Y-%m-%d")
    return str(v)[:10]

def clean_text(text):
    if not text:
        return ""
    text = text.replace("\r", "\n")
    text = text.replace("\u3000", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()

def search_first(text, patterns):
    for p in patterns:
        m = re.search(p, text, re.S)
        if m:
            return clean_text(m.group(1))
    return ""

def normalize_date(s):
    s = clean_text(s)
    s = s.replace("（北京时间）", "").replace("(北京时间)", "")
    s = s.replace("法定节假日除外", "")
    s = re.sub(r"\s+", " ", s)
    s = s.strip(" ，。；;：:")
    return s

def extract_budget(text):
    patterns = [
        r"预算金额[（(]元[）)]?[：:\s]*([^\n；;。]{1,120})",
        r"预算金额[：:\s]*([^\n；;。]{1,120})",
        r"项目预算[：:\s]*([^\n；;。]{1,120})",
        r"采购预算[：:\s]*([^\n；;。]{1,120})",
    ]
    return search_first(text, patterns)

def extract_deadline(text):
    patterns = [
        r"提交投标文件截止时间[：:\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日\s*[0-9]{1,2}:[0-9]{2})",
        r"投标截止时间[：:\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日\s*[0-9]{1,2}:[0-9]{2})",
        r"响应文件提交截止时间[：:\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日\s*[0-9]{1,2}:[0-9]{2})",
        r"开标时间[：:\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日\s*[0-9]{1,2}:[0-9]{2})",
    ]
    return normalize_date(search_first(text, patterns))

def extract_file_time(text):
    """提取报名时间 - 优化版"""
    # 先尝试匹配时间段格式
    time_range_patterns = [
        r"获取.*?文件.*?时间[：:\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日[\s至到~]+[0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日[^\n]{0,100})",
        r"报名.*?时间[：:\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日[\s至到~]+[0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日[^\n]{0,100})",
    ]
    for p in time_range_patterns:
        m = re.search(p, text, re.S)
        if m:
            return normalize_date(m.group(1))
    
    # 再尝试单行匹配
    patterns = [
        r"获取采购文件时间[：:\s]*([^\n]{1,300})",
        r"获取招标文件时间[：:\s]*([^\n]{1,300})",
        r"采购文件获取时间[：:\s]*([^\n]{1,300})",
        r"招标文件获取时间[：:\s]*([^\n]{1,300})",
        r"报名时间[：:\s]*([^\n]{1,300})",
        r"时间[：:\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日.*?)(?:\n|获取|$)",
    ]
    return normalize_date(search_first(text, patterns))

def extract_requirements(text):
    patterns = [
        r"申请人的资格要求[：:\s]*([\s\S]{0,3000}?)(?:\n\s*[三四五六七八九十]+、|\Z)",
        r"投标人资格要求[：:\s]*([\s\S]{0,3000}?)(?:\n\s*[三四五六七八九十]+、|\Z)",
        r"供应商资格要求[：:\s]*([\s\S]{0,3000}?)(?:\n\s*[三四五六七八九十]+、|\Z)",
    ]
    block = search_first(text, patterns)
    if not block:
        return ""
    lines = [clean_text(x) for x in block.split("\n") if clean_text(x)]
    cleaned = []
    for line in lines:
        if line in {"申请人的资格要求", "投标人资格要求", "供应商资格要求"}:
            continue
        cleaned.append(line)
    return "\n".join(cleaned[:10]).strip()

def extract_contacts(text):
    patterns = [
        r"采购人信息[：:\s]*([\s\S]{0,600}?)(?:\n\s*采购代理机构信息|\n\s*项目联系方式|\Z)",
        r"采购人[：:\s]*名称[：:\s]*([^\n]{1,100})",
        r"采购人[：:\s]*([^\n]{1,100})",
    ]
    return search_first(text, patterns)

async def get_detail_text_from_page(page, url):
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=60000)
        await page.wait_for_timeout(1500)

        try:
            await page.wait_for_selector("iframe.content-container-mapFrame", timeout=15000)
            frame_locator = page.frame_locator("iframe.content-container-mapFrame")
            body = frame_locator.locator("body")
            await body.wait_for(timeout=20000)
            await page.wait_for_timeout(1200)
            text = await body.inner_text(timeout=20000)
            if text:
                return clean_text(text)
        except:
            pass

        try:
            text = await page.locator("body").inner_text(timeout=10000)
            return clean_text(text)
        except:
            pass
    except Exception as e:
        print(f"    获取详情页失败: {e}")
    
    return ""

def fetch_list_page(page_no):
    payload = {
        "pageNo": page_no,
        "pageSize": PAGE_SIZE,
        "categoryCode": CATEGORY_CODE,
        "_t": int(datetime.now().timestamp() * 1000),
    }
    resp = requests.post(API_URL, json=payload, headers=HEADERS, timeout=20)
    resp.raise_for_status()
    return resp.json()

def extract_list_items(data):
    try:
        raw_items = data["result"]["data"]["data"]
    except:
        return []
    
    results = []
    for item in raw_items:
        title = str(item.get("title") or "").strip()
        article_id = str(item.get("articleId") or "").strip()
        date = format_date_from_ts(item.get("publishDate") or item.get("publishTime"))
        
        if not title or not article_id or not date:
            continue
            
        detail_url = f"{BASE_URL}/site/detail?parentId={PARENT_ID}&articleId={article_id}"
        results.append({
            '网站名': '上海政府采购网',
            '公告日期': date,
            '报名日期': '',
            '投标截止日期': '',
            '项目名称': title,
            '预算': '',
            '投标要求': '',
            '采购人/代理机构及联系方法': '',
            '具体信息': detail_url,
        })
    return results

async def main():
    print("=" * 60)
    print("上海政府采购公告爬虫 v15 - 格式优化版")
    print("=" * 60)
    
    # 第一步：获取列表数据
    print("\n第一步：获取列表数据...")
    all_items = []
    seen = set()
    page = 1
    
    while True:
        try:
            print(f"  获取第 {page} 页...")
            data = fetch_list_page(page)
            items = extract_list_items(data)
            
            if not items:
                break
            
            added = 0
            for item in items:
                key = item['项目名称'] + item['公告日期']
                if key not in seen and '2026-03-21' <= item['公告日期'] <= '2026-03-29':
                    seen.add(key)
                    all_items.append(item)
                    added += 1
            
            print(f"  第 {page} 页新增 {added} 条，累计 {len(all_items)} 条")
            
            dates = [item['公告日期'] for item in items]
            if dates and all(d < '2026-03-21' for d in dates if d):
                break
            
            page += 1
            time.sleep(0.3)
        except Exception as e:
            break
    
    print(f"\n共获取 {len(all_items)} 条3月21-29日的项目")
    
    # 第二步：筛选弱电工程
    print("\n第二步：筛选弱电工程...")
    weak_items = []
    for item in all_items:
        is_weak, keywords = is_weak_electric(item['项目名称'])
        if is_weak:
            item['匹配关键词'] = ', '.join(set(keywords))
            weak_items.append(item)
    
    print(f"弱电工程类项目: {len(weak_items)} 个")
    
    # 第三步：获取弱电工程详情
    print("\n第三步：获取弱电工程项目详情...")
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()
        
        for idx, item in enumerate(weak_items, 1):
            print(f"  [{idx}/{len(weak_items)}] {item['项目名称'][:40]}...")
            
            text = await get_detail_text_from_page(page, item['具体信息'])
            
            if text:
                item['预算'] = extract_budget(text)
                item['投标截止日期'] = extract_deadline(text)
                item['报名日期'] = extract_file_time(text)
                item['投标要求'] = extract_requirements(text)
                item['采购人/代理机构及联系方法'] = extract_contacts(text)
            
            time.sleep(0.3)
        
        await browser.close()
    
    # 添加序号
    for idx, item in enumerate(weak_items, 1):
        item['序号'] = idx
    
    # 保存 Excel - 优化格式
    print("\n第四步：保存 Excel（优化格式）...")
    
    total_count = len(all_items)
    weak_count = len(weak_items)
    date_range = "2026-03-21 至 2026-03-29"
    
    output_file = '/Users/cui/.openclaw/workspace/上海采购公告_弱电工程类_完整版_v15_格式优化.xlsx'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "弱电工程类项目"
    
    # 优化列宽 - A-G 列重点优化
    column_widths = {
        'A': 6,   # 序号 - 窄一点
        'B': 12,  # 网站名
        'C': 12,  # 公告日期
        'D': 35,  # 报名日期 - 加宽，因为可能包含时间段
        'E': 25,  # 投标截止日期
        'F': 55,  # 项目名称 - 最宽
        'G': 25,  # 预算
        'H': 45,  # 投标要求
        'I': 40,  # 采购人
        'J': 25,  # 匹配关键词
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    link_font = Font(size=10, color="0563C1", underline="single")
    info_font = Font(bold=True, size=11, color="000080")
    normal_font = Font(size=10)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 第一行：统计信息
    info_text = f"日期区间: {date_range} | 总项目数: {total_count} | 弱电工程类项目数: {weak_count}"
    ws.merge_cells('A1:J1')
    info_cell = ws['A1']
    info_cell.value = info_text
    info_cell.font = info_font
    info_cell.alignment = Alignment(horizontal='center', vertical='center')
    info_cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    ws.row_dimensions[1].height = 25
    
    # 标题行
    headers = ['序号', '网站名', '公告日期', '报名日期', '投标截止日期', '项目名称', 
               '预算', '投标要求', '采购人/代理机构及联系方法', '匹配关键词']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # 数据行 - 优化对齐方式
    for row_idx, item in enumerate(weak_items, 3):
        # 序号 - 居中
        cell = ws.cell(row=row_idx, column=1, value=item['序号'])
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = normal_font
        
        # 网站名
        cell = ws.cell(row=row_idx, column=2, value=item['网站名'])
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.font = normal_font
        
        # 公告日期
        cell = ws.cell(row=row_idx, column=3, value=item['公告日期'])
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = normal_font
        
        # 报名日期 - 左对齐，自动换行
        cell = ws.cell(row=row_idx, column=4, value=item['报名日期'])
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.font = normal_font
        
        # 投标截止日期
        cell = ws.cell(row=row_idx, column=5, value=item['投标截止日期'])
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = normal_font
        
        # 项目名称 - 带超链接，左对齐
        name_cell = ws.cell(row=row_idx, column=6, value=item['项目名称'])
        name_cell.hyperlink = item['具体信息']
        name_cell.font = link_font
        name_cell.border = thin_border
        name_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 预算
        cell = ws.cell(row=row_idx, column=7, value=item['预算'])
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.font = normal_font
        
        # 投标要求 - 左对齐，顶部对齐，自动换行
        req_cell = ws.cell(row=row_idx, column=8, value=item['投标要求'])
        req_cell.border = thin_border
        req_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        req_cell.font = normal_font
        
        # 采购人
        cell = ws.cell(row=row_idx, column=9, value=item['采购人/代理机构及联系方法'])
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.font = normal_font
        
        # 匹配关键词
        cell = ws.cell(row=row_idx, column=10, value=item['匹配关键词'])
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.font = normal_font
    
    # 设置所有数据行的默认高度
    for row in range(3, len(weak_items) + 3):
        ws.row_dimensions[row].height = 60  # 默认高度，内容会自动撑开
    
    ws.freeze_panes = 'A3'
    wb.save(output_file)
    
    print(f"\n已保存: {output_file}")
    print(f"\n弱电工程类项目 ({weak_count}个)")

if __name__ == '__main__':
    asyncio.run(main())
