#!/usr/bin/env python3
"""
上海政府采购网爬虫 - 完整版（带详情抓取和筛选）
基于 grab_openapi_v3.py 修改
"""

import os
import sys
import re
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

# API配置
API_URL = "https://www.zfcg.sh.gov.cn/portal/category"
BASE_URL = "https://www.zfcg.sh.gov.cn"
CATEGORY_CODE = "ZcyAnnouncement2"
PARENT_ID = "137027"
SITE_NAME = "上海市政府采购网"
PAGE_SIZE = 15

# 筛选关键词
FILTER_KEYWORDS = ["弱电", "信息化", "智慧校园", "智能化", "安防", "监控", "网络", "机房", "数据中心", "多媒体", "教室", "校园", "教育信息化"]

HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Content-Type": "application/json;charset=UTF-8",
    "Referer": "https://www.zfcg.sh.gov.cn/",
    "Origin": "https://www.zfcg.sh.gov.cn",
}


def format_date_from_ts(v):
    """格式化时间戳"""
    if v is None or v == "":
        return ""
    if isinstance(v, int):
        s = str(v)
        if len(s) == 13:
            return datetime.fromtimestamp(v / 1000).strftime("%Y-%m-%d")
        if len(s) == 10:
            return datetime.fromtimestamp(v).strftime("%Y-%m-%d")
    return str(v)[:10]


def fetch_list_page(page_no: int) -> dict:
    """获取列表页数据"""
    payload = {
        "pageNo": page_no,
        "pageSize": PAGE_SIZE,
        "categoryCode": CATEGORY_CODE,
        "_t": int(datetime.now().timestamp() * 1000),
    }
    resp = requests.post(API_URL, json=payload, headers=HEADERS, timeout=20)
    resp.raise_for_status()
    return resp.json()


def extract_list_items(data: dict) -> list[dict]:
    """提取列表项"""
    try:
        raw_items = data["result"]["data"]["data"]
    except Exception:
        print("接口返回结构异常：")
        print(data)
        return []

    results = []
    for item in raw_items:
        title = str(item.get("title") or "").strip()
        article_id = str(item.get("articleId") or "").strip()
        date = format_date_from_ts(item.get("publishDate") or item.get("publishTime"))
        if not title or not article_id or not date:
            continue
        detail_url = f"{BASE_URL}/site/detail?parentId={PARENT_ID}&articleId={article_id}"
        results.append({
            "项目名称": title,
            "具体信息": detail_url,
            "公告日期": date,
        })
    return results


def filter_items(items: list[dict]) -> list[dict]:
    """根据关键词筛选项目"""
    filtered = []
    for item in items:
        title = item.get("项目名称", "")
        for keyword in FILTER_KEYWORDS:
            if keyword in title:
                item["匹配关键词"] = keyword
                filtered.append(item)
                break
    return filtered


def fetch_by_date(target_date: str = None) -> tuple[str, list[dict]]:
    """获取指定日期的列表"""
    all_items = []
    seen = set()
    base_date = target_date
    page = 1
    max_pages = 10

    while page <= max_pages:
        data = fetch_list_page(page)
        items = extract_list_items(data)
        if not items:
            print("没有数据，停止")
            break

        if base_date is None:
            base_date = items[0]["公告日期"]
            print(f"目标日期: {base_date}")

        added = 0
        stop = False
        for item in items:
            if item["公告日期"] != base_date:
                if added > 0 or len(all_items) > 0:
                    stop = True
                    break
                continue
            
            key = (item["项目名称"], item["具体信息"])
            if key in seen:
                continue
            seen.add(key)
            all_items.append(item)
            added += 1

        print(f"第{page}页新增 {added} 条，累计 {len(all_items)} 条")
        
        if stop:
            print(f"日期变化，停止")
            break
        
        if added == 0 and len(all_items) > 0:
            print("当前页无目标日期数据，停止")
            break
            
        page += 1

    if not all_items:
        raise RuntimeError(f"没有抓到 {base_date} 的列表数据")
    return base_date, all_items


# ---------------- 详情页抓取（简化版） ----------------
def clean_text(text: str) -> str:
    """清理文本"""
    if not text:
        return ""
    text = text.replace("\r", "\n")
    text = text.replace("\u3000", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def extract_budget(text: str) -> str:
    """提取预算金额"""
    patterns = [
        r"预算金额[（(]元[）)]?[：:\s]*([^\n；;。]{1,120})",
        r"预算金额[：:\s]*([^\n；;。]{1,120})",
        r"项目预算[：:\s]*([^\n；;。]{1,120})",
    ]
    for p in patterns:
        m = re.search(p, text)
        if m:
            return clean_text(m.group(1))
    return ""


def extract_file_time(text: str) -> str:
    """提取报名日期（获取招标文件时间）"""
    patterns = [
        r"获取招标文件[\s\S]{0,200}?时间[：:\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日至[0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日)",
        r"获取采购文件[\s\S]{0,200}?时间[：:\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日至[0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日)",
        r"时间[：:\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日至[0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日)[\s\S]{0,100}?获取招标",
    ]
    for p in patterns:
        m = re.search(p, text, re.S)
        if m:
            return clean_text(m.group(1))
    return ""


def extract_deadline(text: str) -> str:
    """提取投标截止日期"""
    patterns = [
        r"提交投标文件截止时间[：:\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日\s*[0-9]{1,2}:[0-9]{2})",
        r"投标截止时间[：:\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日\s*[0-9]{1,2}:[0-9]{2})",
    ]
    for p in patterns:
        m = re.search(p, text)
        if m:
            return clean_text(m.group(1))
    return ""


def extract_requirements(text: str) -> str:
    """提取投标要求"""
    patterns = [
        r"申请人的资格要求[：:\s]*([\s\S]{0,2000}?)(?:\n\s*获取采购文件|\Z)",
        r"投标人资格要求[：:\s]*([\s\S]{0,2000}?)(?:\n\s*获取招标文件|\Z)",
    ]
    for p in patterns:
        m = re.search(p, text, re.S)
        if m:
            return clean_text(m.group(1))[:500]  # 限制长度
    return ""


def extract_contacts(text: str) -> str:
    """提取联系方式"""
    patterns = [
        r"采购人信息[：:\s]*([\s\S]{0,800}?)(?:\n\s*采购代理机构信息|\Z)",
        r"(名\s*称[：：][^\n]+\n地\s*址[：：][^\n]+\n联系方式[：：][^\n]+)",
    ]
    for p in patterns:
        m = re.search(p, text, re.S)
        if m:
            result = clean_text(m.group(1))
            # 去除"附件信息"及之后的内容
            if "附件信息" in result:
                result = result.split("附件信息")[0].strip()
            return result[:300]
    return ""


def enrich_items(items: list[dict]) -> list[dict]:
    """使用Playwright抓取详情页"""
    print("\n正在抓取详情页...")
    
    with sync_playwright() as p:
        # 尝试查找浏览器
        try:
            browser = p.chromium.launch(headless=True)
        except Exception as e:
            print(f"启动浏览器失败: {e}")
            print("将只返回列表数据")
            return items
        
        page = browser.new_page()
        
        for idx, item in enumerate(items, start=1):
            url = item["具体信息"]
            print(f"[{idx}/{len(items)}] {url}")
            
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=60000)
                page.wait_for_timeout(1500)
                
                # 尝试获取iframe内容
                try:
                    frame = page.frame_locator("iframe.content-container-mapFrame")
                    text = clean_text(frame.locator("body").inner_text(timeout=10000))
                except:
                    text = clean_text(page.locator("body").inner_text(timeout=10000))
                
                item["预算"] = extract_budget(text)
                item["报名日期"] = extract_file_time(text)
                item["投标截止日期"] = extract_deadline(text)
                item["投标要求"] = extract_requirements(text)
                item["采购人/代理机构及联系方法"] = extract_contacts(text)
                
            except Exception as e:
                print(f"  抓取失败: {e}")
            
            time.sleep(0.5)
        
        browser.close()
    
    return items


# ---------------- Excel输出 ----------------
def build_dataframe(items: list[dict]) -> pd.DataFrame:
    """构建DataFrame"""
    rows = []
    for i, item in enumerate(items, start=1):
        url = item.get("具体信息", "")
        title = item.get("项目名称", "")
        info_text = f"{title}\n{url}" if title and url else url
        
        rows.append({
            "序号": i,
            "网站名": SITE_NAME,
            "公告日期": item.get("公告日期", ""),
            "报名日期": item.get("报名日期", ""),
            "投标截止日期": item.get("投标截止日期", ""),
            "项目名称": title,
            "预算": item.get("预算", ""),
            "投标要求": item.get("投标要求", ""),
            "采购人/代理机构及联系方法": item.get("采购人/代理机构及联系方法", ""),
            "具体信息": info_text,
            "_具体信息URL": url,
        })
    
    return pd.DataFrame(rows, columns=[
        "序号", "网站名", "公告日期", "报名日期", "投标截止日期",
        "项目名称", "预算", "投标要求", "采购人/代理机构及联系方法",
        "具体信息", "_具体信息URL"
    ])


def beautify_excel(file_name: str):
    """美化Excel格式"""
    wb = load_workbook(file_name)
    ws = wb.active

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 8, "B": 18, "C": 14, "D": 20, "E": 20,
        "F": 46, "G": 18, "H": 50, "I": 42, "J": 58,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # 给"具体信息"列加超链接
    for row_idx in range(2, ws.max_row + 1):
        display_cell = ws[f"J{row_idx}"]
        url_cell = ws[f"K{row_idx}"]
        url = url_cell.value
        if url:
            display_cell.hyperlink = str(url)
            display_cell.style = "Hyperlink"

    for row in ws.iter_rows(min_row=2):
        for cell in row[:10]:
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    ws.column_dimensions["K"].hidden = True
    wb.save(file_name)


def main():
    print("=" * 60)
    print("上海政府采购网爬虫 - 完整版")
    print(f"筛选关键词: {', '.join(FILTER_KEYWORDS)}")
    print("=" * 60)
    
    # 获取当日列表
    print("\n正在获取当日招标公告...")
    base_date, all_items = fetch_by_date()
    print(f"\n当日 ({base_date}) 共 {len(all_items)} 条公告")
    
    # 筛选
    filtered_items = filter_items(all_items)
    print(f"匹配项目: {len(filtered_items)} 条")
    
    if filtered_items:
        print("\n匹配的项目:")
        for item in filtered_items:
            print(f"  - [{item['匹配关键词']}] {item['项目名称']}")
        
        # 抓取详情
        filtered_items = enrich_items(filtered_items)
        
        # 生成Excel
        df = build_dataframe(filtered_items)
        
        desktop_dir = Path.home() / "Desktop"
        target_dir = desktop_dir / "上海招标文件"
        target_dir.mkdir(parents=True, exist_ok=True)

        output = target_dir / f"上海采购公告_筛选_{base_date}.xlsx"
        df.to_excel(output, index=False)
        beautify_excel(str(output))
        
        print(f"\n✓ 已保存: {output}")
        print(f"  匹配条数: {len(df)}")
    else:
        print("\n✗ 没有匹配的项目")
    
    # 保存完整数据
    all_items = enrich_items(all_items)
    df_all = build_dataframe(all_items)
    output_all = Path.home() / "Desktop" / "上海招标文件" / f"上海采购公告_全部_{base_date}.xlsx"
    df_all.to_excel(output_all, index=False)
    beautify_excel(str(output_all))
    print(f"✓ 完整数据已保存: {output_all}")


if __name__ == "__main__":
    main()
