import os
import sys
import glob
import re
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

API_URL = "https://www.zfcg.sh.gov.cn/portal/category"
BASE_URL = "https://www.zfcg.sh.gov.cn"
CATEGORY_CODE = "ZcyAnnouncement2"
PARENT_ID = "137027"
SITE_NAME = "上海市政府采购网"
PAGE_SIZE = 15

HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Content-Type": "application/json;charset=UTF-8",
    "Referer": "https://www.zfcg.sh.gov.cn/",
    "Origin": "https://www.zfcg.sh.gov.cn",
}


# ---------------- 列表抓取 ----------------
def format_date_from_ts(v):
    if v is None or v == "":
        return ""
    if isinstance(v, int):
        s = str(v)
        if len(s) == 13:
            return datetime.fromtimestamp(v / 1000).strftime("%Y-%m-%d")
        if len(s) == 10:
            return datetime.fromtimestamp(v).strftime("%Y-%m-%d")
    return str(v)[:10]


def fetch_list_page(page_no: int) -> dict:
    payload = {
        "pageNo": page_no,
        "pageSize": PAGE_SIZE,
        "categoryCode": CATEGORY_CODE,
        "_t": int(datetime.now().timestamp() * 1000),
    }
    resp = requests.post(API_URL, json=payload, headers=HEADERS, timeout=20)
    resp.raise_for_status()
    return resp.json()


def extract_list_items(data: dict) -> list[dict]:
    try:
        raw_items = data["result"]["data"]["data"]
    except Exception:
        print("接口返回结构异常：")
        print(data)
        return []

    results = []
    for item in raw_items:
        title = str(item.get("title") or "").strip()
        article_id = str(item.get("articleId") or "").strip()
        date = format_date_from_ts(item.get("publishDate") or item.get("publishTime"))
        if not title or not article_id or not date:
            continue
        detail_url = f"{BASE_URL}/site/detail?parentId={PARENT_ID}&articleId={article_id}"
        results.append({
            "项目名称": title,
            "具体信息": detail_url,
            "公告日期": date,
        })
    return results


def fetch_today_list() -> tuple[str, list[dict]]:
    all_items = []
    seen = set()
    base_date = None
    page = 1

    while True:
        data = fetch_list_page(page)
        items = extract_list_items(data)
        if not items:
            print("没有数据，停止")
            break

        if base_date is None:
            base_date = items[0]["公告日期"]
            print("基准日期:", base_date)

        added = 0
        stop = False
        for item in items:
            if item["公告日期"] != base_date:
                stop = True
                break
            key = (item["项目名称"], item["具体信息"])
            if key in seen:
                continue
            seen.add(key)
            all_items.append(item)
            added += 1

        print(f"第{page}页新增 {added} 条，累计 {len(all_items)} 条")
        if stop:
            print("日期变化，停止")
            break
        page += 1

    if not all_items:
        raise RuntimeError("没有抓到列表数据")
    return base_date, all_items


# ---------------- 详情页抓取 ----------------
def clean_text(text: str) -> str:
    if not text:
        return ""
    text = text.replace("\r", "\n")
    text = text.replace("\u3000", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def search_first(text: str, patterns: list[str]) -> str:
    for p in patterns:
        m = re.search(p, text, re.S)
        if m:
            return clean_text(m.group(1))
    return ""


def normalize_cn_datetime_text(s: str) -> str:
    s = clean_text(s)
    s = s.replace("（北京时间）", "").replace("(北京时间)", "")
    s = s.replace("法定节假日除外", "")
    s = re.sub(r"\s+", " ", s)
    s = s.strip(" ，。；;：:")
    return s


def extract_budget(text: str) -> str:
    patterns = [
        r"预算金额[（(]元[）)]?[：:\s]*([^\n；;。]{1,120})",
        r"预算金额[：:\s]*([^\n；;。]{1,120})",
        r"项目预算[：:\s]*([^\n；;。]{1,120})",
        r"采购预算[：:\s]*([^\n；;。]{1,120})",
        r"预算总金额[：:\s]*([^\n；;。]{1,120})",
    ]
    return search_first(text, patterns)


def extract_project_name(text: str, fallback_title: str = "") -> str:
    # 优先保留列表页的完整标题，不再用正文里的简化“项目名称”覆盖它
    fallback_title = clean_text(fallback_title)
    if fallback_title:
        return fallback_title
    return search_first(text, [r"项目名称[：:\s]*([^\n]{1,300})"])


def extract_requirements_full(text: str) -> str:
    block_patterns = [
        r"申请人的资格要求[：:\s]*([\s\S]{0,4000}?)(?:\n\s*[三四五六七八九十]+、\s*获取采购文件|\n\s*三、\s*获取采购文件|\n\s*获取采购文件|\n\s*获取招标文件|\n\s*四、\s*响应文件提交|\n\s*提交投标文件截止时间|\n\s*投标截止时间|\Z)",
        r"投标人资格要求[：:\s]*([\s\S]{0,4000}?)(?:\n\s*[三四五六七八九十]+、\s*获取采购文件|\n\s*三、\s*获取采购文件|\n\s*获取采购文件|\n\s*获取招标文件|\n\s*四、\s*响应文件提交|\n\s*提交投标文件截止时间|\n\s*投标截止时间|\Z)",
        r"供应商资格要求[：:\s]*([\s\S]{0,4000}?)(?:\n\s*[三四五六七八九十]+、\s*获取采购文件|\n\s*三、\s*获取采购文件|\n\s*获取采购文件|\n\s*获取招标文件|\n\s*四、\s*响应文件提交|\n\s*提交投标文件截止时间|\n\s*投标截止时间|\Z)",
    ]
    block = search_first(text, block_patterns)
    if not block:
        return ""
    lines = [clean_text(x) for x in block.split("\n") if clean_text(x)]
    cleaned = []
    for line in lines:
        if line in {"申请人的资格要求", "投标人资格要求", "供应商资格要求"}:
            continue
        cleaned.append(line)
    return "\n".join(cleaned).strip()


def extract_file_time(text: str) -> str:
    block_patterns = [
        r"(?:三、\s*)?获取采购文件([\s\S]{0,800}?)(?:\n\s*四、\s*响应文件提交|\n\s*提交投标文件截止时间|\n\s*投标截止时间|\Z)",
        r"(?:三、\s*)?获取招标文件([\s\S]{0,800}?)(?:\n\s*四、\s*响应文件提交|\n\s*提交投标文件截止时间|\n\s*投标截止时间|\Z)",
    ]
    for p in block_patterns:
        m = re.search(p, text, re.S)
        if m:
            block = clean_text(m.group(1))
            result = search_first(block, [r"时间[：:\s]*([^\n]{1,240})"])
            if result:
                return normalize_cn_datetime_text(result)

    patterns = [
        r"获取采购文件时间[：:\s]*([^\n]{1,240})",
        r"获取招标文件时间[：:\s]*([^\n]{1,240})",
        r"采购文件获取时间[：:\s]*([^\n]{1,240})",
    ]
    return normalize_cn_datetime_text(search_first(text, patterns))


def extract_deadline(text: str) -> str:
    # 先精确匹配“提交投标文件截止时间：2026年04月17日 10:00（北京时间）”这类格式
    direct_patterns = [
        r"提交投标文件截止时间[：:\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日\s*[0-9]{1,2}:[0-9]{2}(?:\s*[（(]北京时间[）)])?)",
        r"投标截止时间[：:\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日\s*[0-9]{1,2}:[0-9]{2}(?:\s*[（(]北京时间[）)])?)",
        r"响应文件提交截止时间[：:\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日\s*[0-9]{1,2}:[0-9]{2}(?:\s*[（(]北京时间[）)])?)",
        r"响应截止时间[：:\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日\s*[0-9]{1,2}:[0-9]{2}(?:\s*[（(]北京时间[）)])?)",
        r"截止时间[：:\s]*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日\s*[0-9]{1,2}:[0-9]{2}(?:\s*[（(]北京时间[）)])?)",
    ]
    result = search_first(text, direct_patterns)
    if result:
        return normalize_cn_datetime_text(result)

    # 再匹配大段标题块
    block_patterns = [
        r"(?:四、\s*)?提交投标文件截止时间、开标时间和地点([\s\S]{0,800}?)(?:\n\s*五、|\n\s*六、|\Z)",
        r"(?:四、\s*)?响应文件提交([\s\S]{0,800}?)(?:\n\s*五、\s*响应文件开启|\n\s*五、\s*开标|\Z)",
        r"(?:四、\s*)?提交投标文件截止时间([\s\S]{0,300}?)(?:\n|\Z)",
    ]
    for p in block_patterns:
        m = re.search(p, text, re.S)
        if m:
            block = clean_text(m.group(1))
            result = search_first(block, [
                r"提交投标文件截止时间[：:\s]*([^\n]{1,240})",
                r"投标截止时间[：:\s]*([^\n]{1,240})",
                r"截止时间[：:\s]*([^\n]{1,240})",
                r"提交截止时间[：:\s]*([^\n]{1,240})",
            ])
            if result:
                return normalize_cn_datetime_text(result)

    # 最后兜底
    fallback_patterns = [
        r"投标截止时间[：:\s]*([^\n]{1,240})",
        r"提交投标文件截止时间[：:\s]*([^\n]{1,240})",
        r"响应文件提交截止时间[：:\s]*([^\n]{1,240})",
        r"响应截止时间[：:\s]*([^\n]{1,240})",
        r"截止时间[：:\s]*([^\n]{1,240})",
    ]
    return normalize_cn_datetime_text(search_first(text, fallback_patterns))


def extract_contacts(text: str) -> str:
    patterns = [
        r"采购人信息[：:\s]*([\s\S]{0,800}?)(?:\n\s*采购代理机构信息|\n\s*项目联系方式|\n\s*申请人的资格要求|\n\s*获取采购文件|\n\s*获取招标文件|\Z)",
        r"采购代理机构信息[：:\s]*([\s\S]{0,800}?)(?:\n\s*项目联系方式|\n\s*申请人的资格要求|\n\s*获取采购文件|\n\s*获取招标文件|\Z)",
        r"项目联系方式[：:\s]*([\s\S]{0,500}?)(?:\n\s*申请人的资格要求|\n\s*获取采购文件|\n\s*获取招标文件|\Z)",
        r"采购人[：:\s]*([\s\S]{0,400}?联系电话[：:\s]*[^\n]{1,120})",
    ]
    return search_first(text, patterns)


def get_detail_text_from_page(page, url: str) -> tuple[str, str]:
    page.goto(url, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(1500)

    page_title = ""
    try:
        page_title = clean_text(page.title())
    except Exception:
        page_title = ""

    try:
        page.wait_for_selector("iframe.content-container-mapFrame", timeout=15000)
        frame_locator = page.frame_locator("iframe.content-container-mapFrame")
        body = frame_locator.locator("body")
        body.wait_for(timeout=20000)
        page.wait_for_timeout(1200)
        text = clean_text(body.inner_text(timeout=20000))
        if text:
            return text, page_title
    except Exception:
        pass

    try:
        text = clean_text(page.locator("body").inner_text(timeout=10000))
        return text, page_title
    except Exception:
        return "", page_title

def resolve_browser_executable() -> str:
    """
    exe 运行时：从 exe 同目录下的 ms-playwright 查找浏览器
    py 运行时：从当前工作目录下的 ms-playwright 查找浏览器
    """
    if getattr(sys, "frozen", False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.getcwd()

    ms_dir = os.path.join(base_dir, "ms-playwright")

    headless_candidates = glob.glob(
        os.path.join(
            ms_dir,
            "chromium_headless_shell-*",
            "chrome-headless-shell-win64",
            "chrome-headless-shell.exe",
        )
    )

    chromium_candidates = glob.glob(
        os.path.join(
            ms_dir,
            "chromium-*",
            "chrome-win64",
            "chrome.exe",
        )
    )

    if headless_candidates:
        return headless_candidates[0]
    if chromium_candidates:
        return chromium_candidates[0]

    raise FileNotFoundError(
        f"未找到 Playwright 浏览器，请确认程序同目录下存在 ms-playwright 文件夹。当前查找路径：{ms_dir}"
    )

def enrich_items(items: list[dict]) -> list[dict]:
    browser_executable = resolve_browser_executable()
    print(f"使用浏览器: {browser_executable}")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            executable_path=browser_executable
        )
        page = browser.new_page()

        for idx, item in enumerate(items, start=1):
            url = item["具体信息"]
            print(f"[{idx}/{len(items)}] {url}")
            try:
                text, page_title = get_detail_text_from_page(page, url)
                item["网页名称"] = page_title or item.get("项目名称", "")
                item["预算"] = extract_budget(text)
                item["项目名称"] = extract_project_name(text, item["项目名称"])
                item["投标要求"] = extract_requirements_full(text)
                item["报名日期"] = extract_file_time(text)
                item["投标截止日期"] = extract_deadline(text)
                item["采购人/代理机构及联系方法"] = extract_contacts(text)
                item["抓取状态"] = "成功" if text else "正文为空"
            except PlaywrightTimeoutError:
                item["网页名称"] = item.get("项目名称", "")
                item["预算"] = ""
                item["投标要求"] = ""
                item["报名日期"] = ""
                item["投标截止日期"] = ""
                item["采购人/代理机构及联系方法"] = ""
                item["抓取状态"] = "超时"
            except Exception as e:
                item["网页名称"] = item.get("项目名称", "")
                item["预算"] = ""
                item["投标要求"] = ""
                item["报名日期"] = ""
                item["投标截止日期"] = ""
                item["采购人/代理机构及联系方法"] = ""
                item["抓取状态"] = f"失败: {str(e)[:80]}"
            time.sleep(0.35)

        browser.close()

    return items


# ---------------- Excel 输出 ----------------
def build_final_dataframe(items: list[dict]) -> pd.DataFrame:
    rows = []
    for i, item in enumerate(items, start=1):
        url = item.get("具体信息", "")
        page_title = clean_text(item.get("网页名称", "")) or clean_text(item.get("项目名称", ""))
        info_text = f"{page_title}\n{url}" if page_title and url else url

        rows.append({
            "序号": i,
            "网站名": SITE_NAME,
            "公告日期": item.get("公告日期", ""),
            "报名日期": item.get("报名日期", ""),
            "投标截止日期": item.get("投标截止日期", ""),
            "项目名称": item.get("项目名称", ""),
            "预算": item.get("预算", ""),
            "投标要求": item.get("投标要求", ""),
            "采购人/代理机构及联系方法": item.get("采购人/代理机构及联系方法", ""),
            "具体信息": info_text,
            "_具体信息URL": url,
        })

    return pd.DataFrame(rows, columns=[
        "序号",
        "网站名",
        "公告日期",
        "报名日期",
        "投标截止日期",
        "项目名称",
        "预算",
        "投标要求",
        "采购人/代理机构及联系方法",
        "具体信息",
        "_具体信息URL",
    ])


def beautify_excel(file_name: str):
    wb = load_workbook(file_name)
    ws = wb.active

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 8,
        "B": 18,
        "C": 14,
        "D": 34,
        "E": 26,
        "F": 46,
        "G": 18,
        "H": 70,
        "I": 42,
        "J": 58,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # 给“具体信息”列加超链接
    # 当前列顺序里：J=具体信息，K=_具体信息URL（隐藏辅助列）
    for row_idx in range(2, ws.max_row + 1):
        display_cell = ws[f"J{row_idx}"]
        url_cell = ws[f"K{row_idx}"]
        url = url_cell.value
        if url:
            display_cell.hyperlink = str(url)
            display_cell.style = "Hyperlink"

    for row in ws.iter_rows(min_row=2):
        max_lines = 1
        for cell in row[:10]:
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            txt = "" if cell.value is None else str(cell.value)
            lines = txt.count("\n") + 1
            approx_lines = max(lines, len(txt) // 28 + 1)
            if approx_lines > max_lines:
                max_lines = approx_lines
        ws.row_dimensions[row[0].row].height = min(max(22, max_lines * 15), 120)

    # 隐藏辅助列 K
    ws.column_dimensions["K"].hidden = True

    wb.save(file_name)


def main():
    base_date, items = fetch_today_list()
    items = enrich_items(items)
    df = build_final_dataframe(items)

    # 导出前不删除辅助列，方便 beautify_excel 给 J 列写入超链接
    desktop_dir = Path.home() / "Desktop"
    target_dir = desktop_dir / "上海招标文件"
    target_dir.mkdir(parents=True, exist_ok=True)

    output = target_dir / f"上海采购公告_整合_{base_date}.xlsx"
    df.to_excel(output, index=False)
    beautify_excel(str(output))
    print(f"完成: {output}")
    print(f"总条数: {len(df)}")


if __name__ == "__main__":
    main()